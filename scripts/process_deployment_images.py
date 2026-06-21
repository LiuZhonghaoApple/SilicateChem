#!/usr/bin/env python3
"""Process website_image_deployment_plan.xlsx → public/images/{page}/ webp + manifest."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
SRC_IMAGES = ROOT / "website_assets_v1" / "images"
PLAN = ROOT / "website_image_deployment_plan.xlsx"
MANIFEST = ROOT / "src" / "content" / "site-images.manifest.json"

MAX_WIDTH = 1920
WEBP_QUALITY = 85

SHEET_FOLDER = {
    "Home": "home",
    "About": "about",
    "Factory": "factory",
    "Products": "products",
    "Export": "export",
    "Certificates": "certifications",
}

SLUG_BY_ORDER = {
    "Products": {
        1: "sodium-metasilicate",
        2: "sodium-metasilicate-granules",
        3: "sodium-metasilicate-granules",
        4: "sodium-metasilicate-anhydrous",
        5: "sodium-metasilicate-anhydrous",
        6: "sodium-metasilicate-pentahydrate",
        7: "sodium-metasilicate-pentahydrate",
        8: "sodium-silicate",
        9: "sodium-silicate",
        10: None,
        11: None,
        12: None,
        13: None,
        14: None,
    }
}


def slugify(text: str) -> str:
    s = text.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "image"


def alt_text(section: str, block: str, page: str) -> str:
    page_label = {
        "/": "Shandong sodium metasilicate factory",
        "/about": "Shandong Zhongzhi Chemical manufacturing site",
        "/factory": "Sodium metasilicate production facility Changyi Shandong",
        "/export": "Sodium metasilicate export packaging and logistics",
        "/export#packaging": "Sodium metasilicate export packaging",
        "/products": "Sodium metasilicate product",
        "/applications": "Sodium metasilicate industrial application",
    }.get(page, "Shandong inorganic silicate manufacturing")

    block_clean = block.replace("_", " ")
    if "Hero" in block or "Slide" in block:
        return f"{page_label} — {block_clean}"
    if "Gallery" in section or "Production" in section:
        return f"{page_label} — {section}: {block_clean}"
    if "Packaging" in section or "Shipping" in section or "Port" in section:
        return f"Sodium metasilicate export — {block_clean} at Changyi Shandong plant"
    return f"{page_label} — {section}: {block_clean}"


def filename_for_row(sheet: str, order: int, section: str, block: str) -> str:
    if sheet == "Home":
        if "Hero" in section:
            return f"hero-{order:02d}.webp"
        if "Factory Strength" in section:
            return "factory-preview.webp"
        return f"production-{order - 4:02d}.webp"
    if sheet == "About":
        names = ["facility-exterior", "production-line", "plant-operations"]
        return f"{names[order - 1]}.webp"
    if sheet == "Factory":
        if "Warehouse" in section:
            return "warehouse-staging.webp"
        return f"line-{order:02d}.webp"
    if sheet == "Products":
        return f"product-{order:02d}.webp"
    if sheet == "Export":
        export_names = [
            "warehouse-staging",
            "bagged-product-01",
            "bagged-product-02",
            "export-loading-01",
            "export-loading-02",
            "export-loading-03",
        ]
        return f"{export_names[order - 1]}.webp"
    return f"{slugify(section)}-{order:02d}.webp"


def to_webp(src: Path, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(src) as im:
        if im.mode in ("RGBA", "P"):
            im = im.convert("RGB")
        w, h = im.size
        if w > MAX_WIDTH:
            nh = int(h * MAX_WIDTH / w)
            im = im.resize((MAX_WIDTH, nh), Image.Resampling.LANCZOS)
        im.save(dest, "WEBP", quality=WEBP_QUALITY, method=6)


def find_source(name: str) -> Path | None:
    for folder in ["01_Factory", "02_Production", "04_Products"]:
        p = ROOT / "lgwjg_image_export" / folder / name
        if p.exists():
            return p
    p = SRC_IMAGES / name
    if p.exists():
        return p
    for f in SRC_IMAGES.iterdir() if SRC_IMAGES.exists() else []:
        if f.name.lower() == name.lower():
            return f
    return None


def load_plan() -> list[dict]:
    wb = load_workbook(PLAN, read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet_name, folder in SHEET_FOLDER.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[3] or row[3] == "—":
                continue
            rows.append(
                {
                    "sheet": sheet_name,
                    "folder": folder,
                    "section": str(row[0]),
                    "block": str(row[1]),
                    "order": int(row[2]),
                    "source": str(row[3]),
                    "page_url": str(row[5]),
                }
            )
    wb.close()
    return rows


def main() -> None:
    rows = load_plan()
    manifest: dict = {
        "home": {"hero": [], "factoryPreview": None, "production": []},
        "about": [],
        "factory": [],
        "products": [],
        "export": {"packaging": [], "shipping": []},
        "certifications": [],
        "bySlug": {},
    }

    for row in rows:
        sheet = row["sheet"]
        folder = row["folder"]
        order = row["order"]
        out_name = filename_for_row(sheet, order, row["section"], row["block"])
        dest = ROOT / "public" / "images" / folder / out_name
        src = find_source(row["source"])
        if not src:
            print(f"[skip] missing: {row['source']}")
            continue
        to_webp(src, dest)
        public = f"/images/{folder}/{out_name}"
        alt = alt_text(row["section"], row["block"], row["page_url"])
        entry = {
            "src": public,
            "alt": alt,
            "section": row["section"],
            "block": row["block"],
            "source": row["source"],
        }
        print(f"  {public}")

        if sheet == "Home":
            if "Hero" in row["section"]:
                manifest["home"]["hero"].append(entry)
            elif "Factory Strength" in row["section"]:
                manifest["home"]["factoryPreview"] = entry
            else:
                manifest["home"]["production"].append(entry)
        elif sheet == "About":
            manifest["about"].append(entry)
        elif sheet == "Factory":
            manifest["factory"].append(entry)
        elif sheet == "Products":
            manifest["products"].append(entry)
            slug_map = SLUG_BY_ORDER.get("Products", {})
            slug = slug_map.get(order)
            if slug:
                role = "hero" if "Hero" in row["block"] else "gallery"
                manifest["bySlug"].setdefault(slug, {})[role] = entry
        elif sheet == "Export":
            if "Packaging" in row["section"]:
                manifest["export"]["packaging"].append(entry)
            else:
                manifest["export"]["shipping"].append(entry)

    # Trust category mapping (factory gallery slots)
    trust_cats = ["production", "lab", "warehouse", "packaging", "shipping"]
    manifest["galleryByCategory"] = {}
    factory = manifest["factory"]
    export_pkg = manifest["export"]["packaging"]
    export_ship = manifest["export"]["shipping"]
    for i, cat in enumerate(trust_cats):
        if i < len(factory):
            manifest["galleryByCategory"][cat] = factory[i]["src"]
        elif i - len(factory) < len(export_pkg):
            manifest["galleryByCategory"][cat] = export_pkg[i - len(factory)]["src"]
        elif export_ship:
            manifest["galleryByCategory"][cat] = export_ship[0]["src"]

    # Ensure empty folders exist
    for folder in SHEET_FOLDER.values():
        (ROOT / "public" / "images" / folder).mkdir(parents=True, exist_ok=True)
        gitkeep = ROOT / "public" / "images" / folder / ".gitkeep"
        if folder == "certifications" and not any((ROOT / "public" / "images" / folder).glob("*.webp")):
            gitkeep.touch()

    MANIFEST.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"\nManifest: {MANIFEST}")
    print(f"Processed: {len(rows)} images")


if __name__ == "__main__":
    main()
