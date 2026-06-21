#!/usr/bin/env python3
"""Process lgwjg export images per Website Placement → public webp assets."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
EXPORT = ROOT / "lgwjg_image_export"
MAPPING = EXPORT / "website_asset_mapping.xlsx"
PUBLIC_FACTORY = ROOT / "public" / "images" / "factory"
PUBLIC_PRODUCTS = ROOT / "public" / "images" / "products"
MANIFEST = ROOT / "src" / "content" / "site-images.manifest.json"

MAX_WIDTH = 1920
WEBP_QUALITY = 85

PLACEMENT_SLUG = {
    "Homepage Hero": "hero",
    "Homepage Factory": "home-factory",
    "Homepage Production": "home-production",
    "Factory Page": "gallery",
    "Product Page": "product",
    "About Us": "about",
    "Laboratory": "lab",
    "Certificate Section": "certificate",
}

FOLDERS = [
    "01_Factory",
    "02_Production",
    "03_Warehouse",
    "04_Products",
    "05_Laboratory",
    "06_Export",
    "07_Team",
    "08_Certificates",
]


def find_source(name: str) -> Path | None:
    for folder in FOLDERS:
        p = EXPORT / folder / name
        if p.exists():
            return p
    return None


def to_webp(src: Path, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(src) as im:
        if im.mode in ("RGBA", "P"):
            im = im.convert("RGB")
        w, h = im.size
        if w > MAX_WIDTH:
            nh = int(h * MAX_WIDTH / w)
            im = im.resize((MAX_WIDTH, nh), Image.Resampling.LANCZOS)
        im.save(dest, "WEBP", quality=WEBP_QUALITY, method=6)


def load_mapping() -> list[dict]:
    wb = load_workbook(MAPPING, read_only=True)
    ws = wb["Website Placement"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(
            {
                "source_name": str(row[0]),
                "category": str(row[1] or ""),
                "placement": str(row[2] or ""),
            }
        )
    wb.close()
    return rows


def main() -> None:
    rows = load_mapping()
    manifest: dict[str, list[str]] = {
        "hero": [],
        "homeFactory": [],
        "homeProduction": [],
        "factoryGallery": [],
        "products": [],
    }

    counters: dict[str, int] = {}

    for row in rows:
        placement = row["placement"]
        slug = PLACEMENT_SLUG.get(placement)
        if not slug:
            continue

        src = find_source(row["source_name"])
        if not src:
            print(f"[skip] missing source: {row['source_name']}")
            continue

        counters[slug] = counters.get(slug, 0) + 1
        n = counters[slug]

        if slug == "product":
            dest = PUBLIC_PRODUCTS / f"product-{n:02d}.webp"
            public_path = f"/images/products/product-{n:02d}.webp"
            manifest["products"].append(public_path)
        elif slug == "gallery":
            dest = PUBLIC_FACTORY / f"gallery-{n:02d}.webp"
            public_path = f"/images/factory/gallery-{n:02d}.webp"
            manifest["factoryGallery"].append(public_path)
        elif slug == "hero":
            dest = PUBLIC_FACTORY / f"hero-{n:02d}.webp"
            public_path = f"/images/factory/hero-{n:02d}.webp"
            manifest["hero"].append(public_path)
        elif slug == "home-factory":
            dest = PUBLIC_FACTORY / f"home-factory-{n:02d}.webp"
            public_path = f"/images/factory/home-factory-{n:02d}.webp"
            manifest["homeFactory"].append(public_path)
        elif slug == "home-production":
            dest = PUBLIC_FACTORY / f"home-production-{n:02d}.webp"
            public_path = f"/images/factory/home-production-{n:02d}.webp"
            manifest["homeProduction"].append(public_path)
        else:
            continue

        to_webp(src, dest)
        print(f"  {dest.relative_to(ROOT)}")

    # Map gallery slots to trust categories (first 5)
    gallery = manifest["factoryGallery"][:5]
    categories = ["production", "lab", "warehouse", "packaging", "shipping"]
    manifest["galleryByCategory"] = {
        cat: gallery[i] if i < len(gallery) else gallery[-1]
        for i, cat in enumerate(categories)
    }

    MANIFEST.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"\nManifest: {MANIFEST}")
    for k, v in manifest.items():
        if isinstance(v, list):
            print(f"  {k}: {len(v)}")


if __name__ == "__main__":
    main()
